import openpyxl as xl
fileName = "Book1.xlxs"
isOpen = False
heading = []
data = []
book = 0
sheet = 0
def createWorkbook():
	global isOpen
	global book
	global sheet
	global data
	global heading
	try:
		if isOpen == False:
			isOpen = True
			book = xl.Workbook()
			sheet = book.active
			while True : 
				head = input("Enter Heading (Exit for exit): ")
				if head.find("Exit") > -1 :
					break
				heading.append(head)
			print("\nAdd Row with (,) Separation or press Exit for Exit")
			print("\n***************************************************")
			for head in heading:
				print(head,end="\t")
			print("\n***************************************************")
			while True:
				row = input("")
				if row.find("Exit") > -1:
					break
				data.append(row.split(","))			
		else:
			choice = input("Do you want to save changes to "+fileName)
			if choice.find("save") > -1 or choice .find("Save"):
				saveWorkbook()

	except Exception as e:
		print("File already in use : ",e)

def openWorkbook():
	try:
		tempName = input("Enter file Name to be Open : ")
		if tempName.find("Exit"):
			return;
		
		sheet = xl.load_workbook(tempName)
		cell = sheet['A1']
		cell = sheet.cell(row,col)
		print(cell.value)
		fileName = tempName

	except Exception as e:
		print("No file exist with ")
		openWorkbook()

def displayWorkbook():
	pass

def editWorkbook():
	pass

def deleteWorkbook():
	pass
def saveWorkbook():
	global fileName
	global heading
	global data
	global isOpen
	global book
	global sheet
	if isOpen == True:
		sheet.append(heading)
		for row in data:
			sheet.append(row)
		fileName = input("Enter File Name : ")
		if fileName.find(".xlsx") == -1:
			fileName = fileName + ".xlsx"		
		book.save(fileName)
	else:
		print("No file opened")
def main():
	global fileName
	while True:
		print("1. Create new Workbook")
		print("2. Open Workbook")
		print("3. Display Workbook Data")
		print("4. Edit Workbook Data")
		print("5. Delete Workbook")
		print("6. Save Workbook")
		print("0. Exit")
		choice = int(input("Enter Choice : "))
		if choice == 1:
			createWorkbook()
		elif choice == 2:
			openWorkbook()
		elif choice == 3:
			displayWorkbook()
		elif choice == 4:
			editWorkbook()
		elif choice == 5:
			deleteWorkbook()
		elif choice == 6:
			saveWorkbook()
		elif choice == 0:
			break
		else:
			print("Invalid Choice")

if __name__ == "__main__":
	main()
	